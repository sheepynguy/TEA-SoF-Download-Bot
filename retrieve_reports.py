import os           # library used to interface with the os
import shutil
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import time
import datetime
import calendar
import openpyxl
from xls2xlsx import XLS2XLSX


# The name of the school. When you add more schools to this list, follow the same format as seen below. The placement of the school in this list should match the destination in the other list.
school_names = [
    "A+ ACADEMY (057829)",
    "GEORGE I SANCHEZ CHARTER (101804)",
    "ADVANTAGE ACADEMY (057806)",
    "ARLINGTON CLASSICS ACADEMY (220802)",
    "CITYSCAPE SCHOOLS (057841)",
    "CUMBERLAND ACADEMY (212801)",
    "GOLDEN RULE CHARTER SCHOOL (057835)",
    "IDEA PUBLIC SCHOOLS (108807)",
    "IMAGINE INTERNATIONAL ACADEMY OF NORTH TEXAS (043801)",
    "LEADERSHIP PREP SCHOOL (061804)",
    "LEGACY PREPARATORY (057846)",
    "LONE STAR LANGUAGE ACADEMY (043802)",
    "MANARA ACADEMY (057844)",
    "MEYERPARK CHARTER (101855)",
    "THE PRO-VISION ACADEMY (101868)",
    "PIONEER TECHNOLOGY & ARTS ACADEMY (057850)",
    "SAN ANTONIO PREPARATORY SCHOOLS (015840)",
    "ST MARY'S ACADEMY CHARTER SCHOOL (013801)",
    "TRINITY BASIN PREPARATORY (057813)",
    "TRIVIUM ACADEMY (061805)",
    "UME PREPARATORY ACADEMY (057845)",
    "VILLAGE TECH SCHOOLS (057847)",
    "WINFREE ACADEMY CHARTER SCHOOLS (057828)",
    "NOVA ACADEMY (057809)",
    "NYOS CHARTER SCHOOL (227804)"
                ]


# The destination for the PDF files. Change these as you like, but make sure the letter casing is exact
folder_paths =  [
    "Gene Zhu - A+ Academy/Financials",
    "Gene Zhu - George I Sanchez AAMA", 
    "Gene Zhu - Advantage Academy/Financials",
    "Gene Zhu - ACA Team Folder/Financials",
    "Gene Zhu - Cityscape Schools",    #
    "Gene Zhu - Cumberland Academy/Financials",
    "Gene Zhu - Golden Rule",
    "Gene Zhu - IDEA Public Schools",
    "Gene Zhu - Imagine",
    "Gene Zhu - Leadership Prep School",
    "Gene Zhu - Legacy Prep Charter Academy",
    "Gene Zhu - Lone Star Language Academy",
    "Gene Zhu - Manara Academy",   #
    "Gene Zhu - Meyerpark Charter",    #
    "Gene Zhu - Pro-vision Academy",   #
    "Gene Zhu - PTAA",
    "Gene Zhu - San Antonio Prep",
    "Gene Zhu - St. Mary's Academy Charter School",    #
    "Gene Zhu - Trinity Basin Prep - TBP",
    "Gene Zhu - Trivium Academy",
    "Gene Zhu - UME Prep",
    "Gene Zhu - Village Tech Schools",
    "Gene Zhu - Winfree Academy Charter Schools"   #
    "Gene Zhu - Nova Academy",
    "Gene Zhu - NYOS Charter School"
]

# The cell location when copying into the master excel file
downloaded_cells = ['N18', 'N19', 'N20', 'N21', 'N22', 'N23', 'N25', 'N26', 'N28', 'N29', 'N30', 'N31', 'N32', 'N33', 'N34', 'N35', 'N37', 'N38', 'N39', 'N40', 'N43', 'N44', 'N45', 'N46', 'N47', 'N49', 'N50', 'N51', 'N52', 'N53', 'N54', 'N55', 'N56', 'N57', 'N58', 'N59', 'N61', 'N62', 'N63', 'N64', 'N65', 'N66', 'N67', 'N68', 'N69', 'N71', 'N72', 'N73', 'N74', 'N77', 'N78', 'N79', 'N81', 'N82', 'N83', 'N84', 'N85', 'N87']




# downloads the files based on the rows listed on the index array
def download_pdf_files(rows, index, school, folder_path):
    # iterates through each row marked by index to be searched and downloaded
    for line in index:
        row = rows[-(line)].find_elements(By.XPATH, ".//td")
        link = row[5].find_element(By.XPATH, ".//a")
        link.click()

        # wait for file to download and rename it, but don't close the window
        old_name = f"C:/Users/{os.getlogin()}/Downloads/report.pdf"
        while not os.path.exists(old_name):
            time.sleep(2)
            
        # get the written out date the file was uploaded in Month Day, Year format
        char_length = len(school)
        month = calendar.month_name[int((row[1].text)[:2].strip("/"))]
        day = ""
        year = ""
        if (row[1].text)[1] == "/":
            day = (row[1].text)[2:4].strip("/")
            year = (row[1].text)[4:9].strip("/ ")
        else:
            day = (row[1].text)[3:5].strip("/")
            year = (row[1].text)[5:10].strip("/ ")

        # rename the file to include the school name and date the report was uploaded
        new_name = school[:char_length - 9] + " SoF " + month + " " + day + ", " + year + ".pdf"
        os.rename(old_name, f"C:/Users/{os.getlogin()}/Downloads/" + new_name)
        # moves the file to a shared and synced onedrive folder
        shutil.move(f"C:/Users/{os.getlogin()}/Downloads/" + new_name, f"C:/Users/{os.getlogin()}/Dynamic Support Solutions/{folder_path}/{new_name}")

    return


# downloads the excel files based on the rows listed, and deposits them into a MASTER excel file
def download_excel_files(rows, index, school):
    for line in index:
        row = rows[-(line)].find_elements(By.XPATH, ".//td")
        link = row[6].find_element(By.XPATH, ".//a")
        link.click()

        old_name = f"C:/Users/{os.getlogin()}/Downloads/report.xls"
        while not os.path.exists(old_name):
            time.sleep(2)

    # convert the fresh downloaded file from xls to xlsx
        old_file = XLS2XLSX(f"C:/Users/{os.getlogin()}/Downloads/report.xls")
        old_file.to_xlsx(f"C:/Users/{os.getlogin()}/Downloads/report.xlsx")

    # take the contents of the report and put it into the MASTER excel file
        wb1 = openpyxl.load_workbook(f"C:/Users/{os.getlogin()}/Downloads/report.xlsx")
        wb2 = openpyxl.load_workbook(f"C:/Users/{os.getlogin()}/Downloads/Master Sof.xlsx")

        ws1 = wb1.active
        ws2 = wb2.active    # this will change when I figure out how the sheets will look
        # make a search for the school sheet in the workbook


        column = "A"
        # retrieve the first empty row in the master sheet
        last_row = 2
        while True:
            if ws2[column + str(last_row)].value is None:
                break
            last_row += 1

        # put the school, date retrieved, and school year
        ws2["A" + str(last_row)].value = school[:len(school) - 9]
        ws2["B" + str(last_row)].value = school[len(school) - 8:].strip("()")
        ws2["C" + str(last_row)].value = row[1].text
        ws2["D" + str(last_row)].value = "2023-2024"

        i = 0
        for col in ws2.iter_cols(min_col=5, max_col=len(downloaded_cells)+4, min_row=last_row, max_row=last_row):
            for cell in col:
                cell.value = ws1[downloaded_cells[i]].value
                cell.number_format = "0.00"
            i += 1

        # save the contents of the data that was just dumped into the file
        wb2.save(f"C:/Users/{os.getlogin()}/Downloads/Master Sof.xlsx")
        wb2.close()
        # delete the reports that we no longer need
        os.remove(f"C:/Users/{os.getlogin()}/Downloads/report.xls")
        os.remove(f"C:/Users/{os.getlogin()}/Downloads/report.xlsx")
    

    return



# creates an instance of the Edge driver
driver = webdriver.Edge()
#  navigates to the TEA page to find the SoF reports
driver.get("https://tealprod.tea.state.tx.us/fsp/Reports/ReportSelection.aspx")

# iterate through every school to pull reports from their table
for j in range(len(school_names)):
    # selects the Summary of Finances for the first drop down
    form_type = Select(driver.find_element(By.ID, "ctl00_Body_ReportTypeDropDownList"))
    form_type.select_by_value('SummaryOfFinance')
    # clicks the submit button to move onto the next two forms
    button = driver.find_element(By.ID, "ctl00_Body_SelectButton")  # use this same id for the next select button
    button.click()

    # selects the school year for the second drop down
    school_year = Select(driver.find_element(By.ID, "ctl00_Body_SchoolYearDropDownList"))
    school_year.select_by_value("2024")

    # inputs the school name and presses submit
    name_input = driver.find_element(By.ID, "ctl00_Body_DistrictIdTextBox")
    name_input.send_keys(school_names[j])
    name_input.send_keys(Keys.ENTER)

    # Grab the latest row and check if it was the last month's report
    table = driver.find_element(By.ID, "ctl00_Body_SofDistrictRunGridView")
    rows = table.find_elements(By.XPATH, ".//tr")


    i = 1
    # for the upcoming school year, you'll need to put in some kind of input so that it knows to switch between different school years
    month_today = datetime.datetime.now().strftime("%m")
    if month_today[0] == "0":
        month_today = month_today.strip("0")

    # get the negative index of the rows that have the previous month
    index = []
    for i in range(len(rows)):
        row = rows[-(i+1)].find_elements(By.XPATH, ".//td")
        month_web = row[1].text
        if month_web[1] == "/":
            month_web = month_web[:1]
        else:
            month_web = month_web[:2]
        
        # if the previous month shows up, then we add it to the index array
        if int(month_today)-1 == int(month_web):
            index.append(i+1)
        elif int(month_today)-1 > int(month_web):
            break

    download_pdf_files(rows, index, school_names[j], folder_paths[j])
    download_excel_files(rows, index, school_names[j])

    # after downloading the reports, return back to the drop down page to repeat the process for the next school
    reset = driver.find_element(By.ID, "ctl00_Body_SofDistrictRunCancelButton")
    reset.click()