import os           # library used to interface with the os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import time
import openpyxl
from xls2xlsx import XLS2XLSX

# The name of the school. When you add more schools to this list, follow the same format as seen below. The placement of the school in this list should match the destination in the other list.
school_names = [
    "A+ ACADEMY (057829)",
    "GEORGE I SANCHEZ CHARTER (101804)",
    "ADVANTAGE ACADEMY (057806)",
    "ARLINGTON CLASSICS ACADEMY (220802)",
    "CITYSCAPE SCHOOLS (057841)",
    "CUMBERLAND ACADEMY (212801)",
    "GOLDEN RULE CHARTER SCHOOL (057835)",
    "IDEA PUBLIC SCHOOLS (108807)",
    "IMAGINE INTERNATIONAL ACADEMY OF NORTH TEXAS (043801)",
    "LEADERSHIP PREP SCHOOL (061804)",
    "LEGACY PREPARATORY (057846)",
    "LONE STAR LANGUAGE ACADEMY (043802)",
    "MANARA ACADEMY (057844)",
    "MEYERPARK CHARTER (101855)",
    "THE PRO-VISION ACADEMY (101868)",
    "PIONEER TECHNOLOGY & ARTS ACADEMY (057850)",
    "SAN ANTONIO PREPARATORY SCHOOLS (015840)",
    "ST MARY'S ACADEMY CHARTER SCHOOL (013801)",
    "TRINITY BASIN PREPARATORY (057813)",
    "TRIVIUM ACADEMY (061805)",
    "UME PREPARATORY ACADEMY (057845)",
    "VILLAGE TECH SCHOOLS (057847)",
    "WINFREE ACADEMY CHARTER SCHOOLS (057828)",
    "NOVA ACADEMY (057809)",
    "NYOS CHARTER SCHOOL (227804)"
                ]


# The destination for the PDF files. Change these as you like, but make sure the letter casing is exact
folder_paths =  [
    "Gene Zhu - A+ Academy/Financials",
    "Gene Zhu - George I Sanchez AAMA", 
    "Gene Zhu - Advantage Academy/Financials",
    "Gene Zhu - ACA Team Folder/Financials",
    "Gene Zhu - Cityscape Schools",    #
    "Gene Zhu - Cumberland Academy/Financials",
    "Gene Zhu - Golden Rule",
    "Gene Zhu - IDEA Public Schools",
    "Gene Zhu - Imagine",
    "Gene Zhu - Leadership Prep School",
    "Gene Zhu - Legacy Prep Charter Academy",
    "Gene Zhu - Lone Star Language Academy",
    "Gene Zhu - Manara Academy",   #
    "Gene Zhu - Meyerpark Charter",    #
    "Gene Zhu - Pro-vision Academy",   #
    "Gene Zhu - PTAA",
    "Gene Zhu - San Antonio Prep",
    "Gene Zhu - St. Mary's Academy Charter School",    #
    "Gene Zhu - Trinity Basin Prep - TBP",
    "Gene Zhu - Trivium Academy",
    "Gene Zhu - UME Prep",
    "Gene Zhu - Village Tech Schools",
    "Gene Zhu - Winfree Academy Charter Schools"   #
    "Gene Zhu - Nova Academy",
    "Gene Zhu - NYOS Charter School"
]

# The cell location when copying into the master excel file
downloaded_cells = ['N18', 'N19', 'N20', 'N21', 'N22', 'N23', 'N25', 'N26', 'N28', 'N29', 'N30', 'N31', 'N32', 'N33', 'N34', 'N35', 'N37', 'N38', 'N39', 'N40', 'N43', 'N44', 'N45', 'N46', 'N47', 'N49', 'N50', 'N51', 'N52', 'N53', 'N54', 'N55', 'N56', 'N57', 'N58', 'N59', 'N61', 'N62', 'N63', 'N64', 'N65', 'N66', 'N67', 'N68', 'N69', 'N71', 'N72', 'N73', 'N74', 'N77', 'N78', 'N79', 'N81', 'N82', 'N83', 'N84', 'N85', 'N87']

# Cell titles for repeated use
cell_titles = ['District Name', 'Charter ID', 'Date Uploaded', 'School Year', 'Refined ADA', 'Reg. Prog. ADA', 'SpEd FTEs', 'CT FTEs', 'Weighted ADA', 'PEIMS Enroll','Prior TY State Cert. Prop. Value', 'Curr. TY State Cert. Prop. Value', 'Curr. TY M&O Tax Rate', 'Curr. TY Tier 1 M&O Tax Rate', 'Max. Compress. Tax Rate', 'SY Tax Collections', 'Curr. TY I&S Tax Rate', 'SY I&X Tax Collect.', 'SY Total Tax Collect.', 'SY Total Tax Levy', 'District Basic Allot * TR/MCR', 'SSA ADA', 'ASF ADA', 'Per Capita Rate', '11-Reg. Prog. Allot 48.051', 'Sm.&Mid-Size Allot 48.101', '23-SpEd Adj. Allot 48.102', '37-Dyslexia Allot 48.103', '24-Compens. Edu Allot 48.104', '25-Bilingual Ed Allot 48.105', '22-CT Allot 48.106', '11-Public Ed Grant 48.107', '36-Early Ed Allot 48.108', '21-GT Adj Allot 48.109', '38-CCMR Outcomes Bonus 48.110', 'Fast Growth Allot 48.111', 'Teacher Incentive Allot 48.112', 'Mentor Prog Allot 48.114', 'School Safety Allot 48.115', 'R-PEP Allot & Outcomes Bonux 48.118', '99-Transport Allot 48.151', '99-New Instruc Facility Allot 48.152', 'Dropout Rec.&Res. Placement Facil. Allot 48.153', 'Tui Allot (Dist. w/ not all GL) 48.154', 'College Prep Assess Reimburse 48.155', 'Cert Exam Reimburse 48.156', 'Total Cost Tier 1', 'Local Fund Assign', 'Per Capita Dist. from ASF', 'FSP State Share Tier 1', 'Tier 2', 'Other Prog', 'Total FSP Op. Funding', '199/5812 - FSF', '199/5811 - ASF', '410/5829 - IM&TF', '599/5829 - EDA', '599/5829 - IFA(Bond)', '199/5829 - IFA(Lease Purchase)', 'ASAHE for Facil', 'Total FSP/ASF State Aid', 'Local Revenue in Excess Ent.']



def download_excel_files(rows, index, school):
    for line in index:
        row = rows[line].find_elements(By.XPATH, ".//td")
        link = row[6].find_element(By.XPATH, ".//a")
        link.click()

        old_name = f"C:/Users/{os.getlogin()}/Downloads/report.xls"
        while not os.path.exists(old_name):
            time.sleep(2)

    # convert the fresh downloaded file from xls to xlsx
        old_file = XLS2XLSX(f"C:/Users/{os.getlogin()}/Downloads/report.xls")
        old_file.to_xlsx(f"C:/Users/{os.getlogin()}/Downloads/report.xlsx")

    # take the contents of the report and put it into the MASTER excel file
        wb1 = openpyxl.load_workbook(f"C:/Users/{os.getlogin()}/Downloads/report.xlsx")
        wb2 = openpyxl.load_workbook(f"C:/Users/{os.getlogin()}/Downloads/Master Sof.xlsx")

        ws1 = wb1.active
        ws2 = wb2.active    # this will change when I figure out how the sheets will look
        # make a search for the school sheet in the workbook


        column = "A"
        # retrieve the first empty row in the master sheet
        last_row = 2
        while True:
            if ws2[column + str(last_row)].value is None:
                break
            last_row += 1

        # put the school, date retrieved, and school year
        ws2["A" + str(last_row)].value = school[:len(school) - 9]
        ws2["B" + str(last_row)].value = school[len(school) - 8:].strip("()")
        ws2["C" + str(last_row)].value = row[1].text
        ws2["D" + str(last_row)].value = "2023-2024"

        i = 0
        for col in ws2.iter_cols(min_col=5, max_col=len(downloaded_cells)+4, min_row=last_row, max_row=last_row):
            for cell in col:
                cell.value = ws1[downloaded_cells[i]].value
                cell.number_format = "0.00"
            i += 1

        # save the contents of the data that was just dumped into the file
        wb2.save(f"C:/Users/{os.getlogin()}/Downloads/Master Sof.xlsx")
        wb2.close()
        # delete the reports that we no longer need
        os.remove(f"C:/Users/{os.getlogin()}/Downloads/report.xls")
        os.remove(f"C:/Users/{os.getlogin()}/Downloads/report.xlsx")
    return




# creates an instance of the Edge driver
driver = webdriver.Chrome()
#  navigates to the TEA page to find the SoF reports
driver.get("https://tealprod.tea.state.tx.us/fsp/Reports/ReportSelection.aspx")


# creates the headers for the master SoF excel file
#old_file = XLS2XLSX(f"C:/Users/{os.getlogin()}/Downloads/Master SoF.xls")      # uncomment these two lines if the Master SoF originally has an .xls extension
#old_file.to_xlsx(f"C:/Users/{os.getlogin()}/Downloads/Master SoF.xlsx")
wb = openpyxl.load_workbook(f"C:/Users/{os.getlogin()}/Downloads/Master Sof.xlsx")
ws = wb.active
max_col = len(cell_titles)
k = 0
for col in ws.iter_cols(min_col = 1, max_col=max_col, min_row=1, max_row=1):
    for cell in col:
        cell.value = cell_titles[k]
    k += 1
wb.save(f"C:/Users/{os.getlogin()}/Downloads/Master Sof.xlsx")
wb.close()


# part of the main function
# iterate through every school to pull reports from their table
for j in range(len(school_names)):
    # selects the Summary of Finances for the first drop down
    form_type = Select(driver.find_element(By.ID, "ctl00_Body_ReportTypeDropDownList"))
    form_type.select_by_value('SummaryOfFinance')
    # clicks the submit button to move onto the next two forms
    button = driver.find_element(By.ID, "ctl00_Body_SelectButton")  # use this same id for the next select button
    button.click()

    # selects the school year for the second drop down
    school_year = Select(driver.find_element(By.ID, "ctl00_Body_SchoolYearDropDownList"))
    school_year.select_by_value("2024")

    # inputs the school name and presses submit
    name_input = driver.find_element(By.ID, "ctl00_Body_DistrictIdTextBox")
    name_input.send_keys(school_names[j])
    name_input.send_keys(Keys.ENTER)


###This section uses the wrong dates that I'm looking at, so I need to be more specific in what I want
    # makes an index list of which reports to download for a multi-month time span
    # does not include the ending_month
    # use <>_year variables to indicate how far back you want to get the reports from
    starting_month = 5
    starting_year = 2024
    ending_month = 9
    ending_year = 2024

    # get the location of the rows
    table = driver.find_element(By.ID, "ctl00_Body_SofDistrictRunGridView")
    rows = table.find_elements(By.XPATH, ".//tr")

    # iterate through the table to find the first report to download
    saved = 0
    index = []
    for i in range(1, len(rows)):
        row = rows[i].find_elements(By.XPATH, ".//td")
        
        # if the year corresponds with that range, then we will start pulling from the 
        date = row[1].text[:10].strip(" ")
        if " " in date:
            ind = date.index(" ")
            date = date[:ind]
        month = row[1].text[:2].strip("/")
        year = date[-4:]
        if int(year) == starting_year and int(month) == starting_month:
            saved = i 
            break

    for i in range(saved, len(rows)):
        row = rows[i].find_elements(By.XPATH, ".//td")

        # get the month and year of the report so we know whether or not to finish 
        date = row[1].text[:10].strip(" ")
        if " " in date:
            ind = date.index(" ")
            date = date[:ind]
        month = row[1].text[:2].strip("/")
        year = date[-4:]
        if month == ending_month and year == ending_year:
            break
        index.append(i)
        

    
    download_excel_files(rows, index, school_names[j])

    reset = driver.find_element(By.ID, "ctl00_Body_SofDistrictRunCancelButton")
    reset.click()


